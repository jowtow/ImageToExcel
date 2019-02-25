import openpyxl
import os
import sys
import datetime
import argparse

from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

from PIL import Image

def getColumnStr(num):
    columnStr = ''
    headerNum = num / 26
    while(headerNum >= 1):
        columnStr = columnStr + chr(int(headerNum%26) + 64)
        headerNum = headerNum / 26
    columnStr = columnStr + chr(num % 26 + 65)
    return columnStr

def normalizeCells(ws,height,width):
    for i in range(width):
        ws.column_dimensions[getColumnStr(i)].width = 1
    for i in range(1,height):
        ws.row_dimensions[i].height = 5

def getSensitivityPixels(pixels,height,width,pixelation):
    newPixels = []
    for x in range(width):
        pixelRow = []
        for y in range(height):
            pixel = ''
            r = []
            g = []
            b = []
            
            for i in range(pixelation):
                x1 = x*pixelation
                for j in range(pixelation):
                    y1 = y*pixelation
                    rgb = pixels[x1 + i, y1 + j] 
                    r.append(rgb[0])
                    g.append(rgb[1])
                    b.append(rgb[2])
            rAvg = int(sum(r) / float(len(r)))
            gAvg = int(sum(g) / float(len(g)))
            bAvg = int(sum(b) / float(len(b)))
            pixelRow.append((rAvg,gAvg,bAvg))
        newPixels.append(pixelRow)
    return newPixels


def getRgbHarshPixels(pixels, harshness):
    h = pow(2,harshness)
    for i in range(len(pixels)):
        for j in range(len(pixels[i])):
                r = round(pixels[i][j][0]/h)*h
                g = round(pixels[i][j][1]/h)*h
                b = round(pixels[i][j][2]/h)*h
                if r > 255:
                    r = 255
                if g > 255:
                    g = 255
                if b > 255:
                    b = 255
                pixels[i][j] = [r,g,b]
    return pixels

def fillCellsWithPixels(ws,pixels,height,width):
    for i in range(1,width):
        print(f'{round(i/(1.0*width)*100,2)}%', end='\r')
        x = getColumnStr(i)
        for y in range(1,height):
            color = pixels[i][y]
            hexColor = '%02x%02x%02x' % tuple(color)
            fill1 = PatternFill(fill_type='solid',start_color=hexColor,end_color=hexColor)
            ws.cell(row=y,column=i).fill = fill1
    print("100%   ", end='\r')


parser = argparse.ArgumentParser(description="ImageToExcel")
parser.add_argument('file', type=str, help="the filepath of image to convert to excel")
parser.add_argument('-p', '--pixelation', type=int, default=1, help="how pixilated you want the image to be")
parser.add_argument('-s','--saturation', type=int, default=0, choices=list(range(0,9)), help="how saturated you want the image to be")
args = parser.parse_args()

#Setup image
im = Image.open(args.file)
pixels = im.load()
width = int(im.size[0] / args.pixelation)
height = int(im.size[1] / args.pixelation)


#Setup worksheet
wb = Workbook()
ws = wb.active
ws.insert_cols(width)
ws.insert_rows(height)


if (args.pixelation >= 1):
    pixels = getSensitivityPixels(pixels,height,width,args.pixelation)
if(args.saturation > 0 ):
    pixels = getRgbHarshPixels(pixels, args.saturation)

normalizeCells(ws,height,width)
fillCellsWithPixels(ws,pixels,height,width)


fileString = datetime.datetime.now().strftime("%m-%d-%Y___%H-%M-%S")
wb.save("results/" + fileString + '.xlsx')